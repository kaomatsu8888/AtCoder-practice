#Excel ファイルを PDF 化する
import sys
import datetime
import openpyxl #openpyxlをインポート 
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.platypus import SimpleDocTemplate,Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
def excelToPdf(excelFile,pdfFile):
    try:
        workBook = openpyxl.load_workbook(excelFile) #エクセルブックを開く
    except FileNotFoundError:
        print("ファイルが見つかりません")
        sys.exit()
    sheet = workBook.active #アクティブなワークシートを選択
    document = SimpleDocTemplate(pdfFile, pagesize=A4)
    #フリーフォント源真ゴシックを利用(http://jikasei.me/font/genshin/からダウンロード)
    pdfmetrics.registerFont(TTFont("GenShinGothic-Normal","GenShinGothic-Normal.ttf"))
    elements = []
    tableData = []
    for row in sheet.rows:#行ループ
        recordList = []
        for cell in row: #列ループ
            if sheet.cell(row = 1, column = cell.column). value == '売上金額' and cell.row != 1: #もしも列が売上金額のとき
                cell.value = "¥{:,d}".format(cell.value)
            if sheet.cell(row = 1, column = cell.column). value == '日付' and cell.row != 1: #もしも列が日付のとき
                cell.value = str(cell.value.year)+"-"+str(cell.value.month)+"-"+str(cell.value.day)
            recordList.append(cell.value)
        tableData.append(recordList)
    table=Table(tableData)
    table.setStyle(TableStyle([ #属性値、開始位置、終了位置、属性
        ("ALIGN", (0,0),(-1,0), "CENTER"), #行見出しを中央寄せにする
        ('BACKGROUND',(0,0),(-1,0),colors.lightgreen), #行見出しをライトグリーンにする
        ("ALIGN", (0,1), (0,-1), "RIGHT"), #日付の列を右寄せにする
        ("ALIGN", (2,1), (2,-1), "RIGHT"), #売上金額の列を右寄せにする
        ('TEXTCOLOR',(0,0),(1,-1),colors.black), #文字色を黒にする
        ('FONT', (0, 0), (-1, -1), "GenShinGothic-Normal", 20), #フォントとフォントサイズを設定
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),])) #ケイ線の設定
    elements.append(table) 
    try:
        document.build(elements) #PDFの書き出し
    except FileNotFoundError:
        print("出力用のフォルダがありません")
        sys.exit()
excelToPdf("売上リスト(pg22_01用).xlsx","売上リストPDF.pdf")