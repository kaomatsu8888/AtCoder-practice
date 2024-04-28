#指定したブック、セル範囲間でコピーアンドペーストするプログラム  応用編
import openpyxl #エクセルのブックを操作するライブラリ の読み込み
import re #文字列操作をするライブラリの読み込み
#コピーアンドペーストする関数
#bookName:対象のエクセルブック名 ,sheetName:シート名,fromCellRange:コピー元セル範囲,toCell:コピー先セル番地 
def copyAndPaste(bookName,sheetName,fromCellRange,toCell):
    workBook = openpyxl.load_workbook(bookName) #指定したファイル名でエクセルブックを開く
    workSheet = workBook[sheetName]
    #指定したシート名のシートを取得する
    fromStartCell ="" #コピー元開始セル
    fromEndCell = "" #コピー元終了セル
    if fromCellRange.find(':') != -1:#:が含まれているかどうか判定 
        workCellData = re.split(":",fromCellRange) 
        fromStartCell = workCellData[0]
        fromEndCell = workCellData[1]
        fromStartCell = getRowAndColumn(fromStartCell)
        fromEndCell = getRowAndColumn(fromEndCell)
    else: #:が含まれていないときに処理される
        fromStartCell = getRowAndColumn(fromCellRange)
        fromEndCell = fromStartCell
    toCell = getRowAndColumn(toCell)
    wrow = 0 #貼り付け行位置計算用の変数
    for i in range(fromStartCell[0],fromEndCell[0]+1): #行ループ
        wcol = 0 #貼り付け列位置計算用の変数
        for j in range(fromStartCell[1],fromEndCell[1]+1): #列ループ
            workSheet.cell(row = toCell[0] + wrow, column = toCell[1] + wcol).value = workSheet.cell(row = i, column = j).value
            wcol = wcol + 1
        wrow = wrow + 1
    workBook.save(bookName) #ブックを保存する
    #指定したセル名から行番号と列番号を取得
def getRowAndColumn(cellAdress): #指定したセル名から行番号と列番号を取得
    column = re.split("[\d]",cellAdress) #列名の取り出し(リスト)
    column = column[0] #列名を取り出し(文字列)
    row = cellAdress.replace(column,"") #行番号を取得
    row = int(row) #行番号を文字列から数値に変換
    column = openpyxl.utils.column_index_from_string(column) #列名を列番号に変換
    return [row,column] #行番号と列番号をリストで戻す

#関数の呼び出し A1からD4の範囲をA10に貼り付け 
copyAndPaste("商品リスト.xlsx","Sheet1","A1:D4","A10")
