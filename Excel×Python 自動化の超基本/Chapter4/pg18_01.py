#複数の動作を1つのプログラムでおこなう
import openpyxl #openpyxlをインポート
from openpyxl.styles import PatternFill, Border,Side, Alignment, Protection, Font
#列見出しのフォントを定義
column_font = Font(name='MS Pゴシック',size=16,bold=False,italic=True,vertAlign=None,underline='none',strike=False,color='000000')
#列見出しの配置を定義
column_alignment=Alignment(horizontal='center',vertical='bottom',text_rotation=0,wrap_text=False,shrink_to_fit=False,indent=0)
#セルのケイ線を定義
border = border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
#数値の書式を定義
number_format = '¥#,##0;¥-#,##0'
workBook = openpyxl.load_workbook("新宿店売上リスト.xlsx") #エクセルブックを開く
sheet = workBook["新宿店"] #シート「新宿店」を選択
sheet.column_dimensions["A"].width = 15 #列の幅を設定(文字数)
sheet.column_dimensions["B"].width = 15 #列の幅を設定(文字数)
sheet.column_dimensions["C"].width = 15 #列の幅を設定(文字数)
for row in sheet: #行ループ
    for cell in row: #列ループ
        if row == 1: #1行目に列見出しの書式を割り当てる
            cell.font = column_font #列見出しのフォントを設定 cell.alignment = sheet["B1"].alignment = sheet["C1"].alignment = column_alignment #列 見出しの配置を設定
        if sheet.cell(row = 1, column = cell.column).value == '売上金額': #もしも列が売上金額のとき
            cell.number_format = number_format #セルに数値形式の書式を設定
        cell.border = border #セルのケイ線を設定
workBook.save('新宿店売上リスト(完成).xlsx') #ファイル名を指定して保存