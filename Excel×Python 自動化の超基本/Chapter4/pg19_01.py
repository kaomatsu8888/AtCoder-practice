#条件を満たす行の色を変更する
import openpyxl #openpyxlをインポート
from openpyxl.styles import PatternFill, Border,Side, Alignment, Protection, Font
#売上金額が35万円以上のセルのフォントを定義 
high_font = Font(name='MS Pゴシック',size=14,bold=True,italic=True,vertAlign=None,underline='none',strike=False,color='FF0000')#数値の書式を定義
number_format = '¥#,##0;¥-#,##0'
#ファイルを開く
workBook = openpyxl.load_workbook("新宿店売上リスト.xlsx") #エクセルブックを開く
sheet = workBook["新宿店"] #シート「新宿店」を選択
for row in sheet: #行ループ
    for cell in row: #列ループ
        if sheet.cell(row = 1, column = cell.column). value == '売上金額': #もしも列が売上金額のとき
            sales_column = cell.column #売上金額のセル位置 を格納
            cell.number_format = number_format #セルに 数値形式の書式を設定
        if cell.row != 1 and sheet.cell(row=cell. row,column= sales_column).value >= 350000:#セ ルの値が35万円以上の時
            cell.font = high_font #セルのフォントを設定
workBook.save('新宿店売上リスト(完成)2.xlsx') #ファイル名を指定して保存
